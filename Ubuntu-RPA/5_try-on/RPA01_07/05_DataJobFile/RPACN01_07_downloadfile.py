import base64
import ctypes
import re
import shutil
import signal
import sys, win32com.client
import traceback

import psutil
import pytz
import win32api, win32gui, win32con, win32ui, time, os, subprocess
import pandas as pd

from datetime import datetime,timedelta
import logging
import wmi
from exchangelib import Configuration, Credentials, Account, FileAttachment, Message, DELEGATE
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import pywintypes
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.chrome.service import Service as ChromeService, Service
from selenium.webdriver.chrome.options import Options as ChromeOptions, Options
from selenium.webdriver.common.keys import Keys
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
import os
import sys

def get_base_path():
    """获取基础路径，兼容exe打包和开发环境"""
    if getattr(sys, 'frozen', False):
        # 打包后的exe运行
        return os.path.dirname(sys.executable)
    else:
        # 开发环境运行
        return os.path.dirname(os.path.abspath(__file__))

def get_project_root():
    base_path = get_base_path()
    return os.path.abspath(os.path.join(base_path, '..'))

def ensure_working_dir():
    os.chdir(get_base_path())

def kill_sap():
    c = wmi.WMI()

    for process in c.Win32_Process(name="saplogon.exe"):
        print(process.ProcessId, process.Name)
        process.Terminate()

def kill_excel_processes():
    for proc in psutil.process_iter():
        try:
            if proc.name() == 'EXCEL.EXE':
                proc.kill()
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

def kill_chrome_processes(retry_count=7):
    attempts = 0
    wmi = win32com.client.GetObject("winmgmts:")
    while attempts < retry_count:
        try:
            for proc in psutil.process_iter(['pid', 'name']):
                if proc.info['name'] == 'chrome.exe':
                    try:
                        proc.terminate()
                        proc.wait(timeout=5)  # 等待进程终止
                        print(f'已经杀死进程 {proc.info["name"]} (PID: {proc.info["pid"]})')
                    except psutil.TimeoutExpired:
                        logger.warning(
                            f"进程 {proc.info['name']} (PID: {proc.info['pid']}) 未在指定时间内终止，强制终止")
                        proc.kill()
                        proc.wait(timeout=5)
                        print(f'已经强制杀死进程 {proc.info["name"]} (PID: {proc.info["pid"]})')
                    except psutil.NoSuchProcess:
                        logger.warning(f"进程 {proc.info['name']} (PID: {proc.info['pid']}) 已经不存在")
                    except psutil.AccessDenied:
                        logger.error(f"没有权限终止进程 {proc.info['name']} (PID: {proc.info['pid']})")
            break
        except Exception as e:
            attempts += 1
            logger.error(f"杀死进程失败{e}")
            if attempts < retry_count:
                logger.info(f"重试 ({attempts}/{retry_count})...")
            else:
                logger.error(f"达到最大登录次数，登录失败{e}")

def logging_info():
    # 创建一个logger
    global logger
    log_directory =Config['log_directory']  # 替换为你需要的具体路径
    log_filename = f"log_{datetime.now().strftime('%Y-%m-%d')}.log"
    log_file_path = os.path.join(log_directory, log_filename)

    # 确保日志文件所在的目录存在
    os.makedirs(log_directory, exist_ok=True)

    # 创建一个logger
    logger = logging.getLogger('my_logger')
    logger.setLevel(logging.DEBUG)  # 设置全局日志级别

    # 创建一个FileHandler，用于写入日志文件
    file_handler = logging.FileHandler(log_file_path, encoding='utf-8')  # 设置编码为 utf-8
    file_handler.setLevel(logging.DEBUG)  # 设置handler的日志级别

    # 定义handler的输出格式，包括时间戳（精确到秒）
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    file_handler.setFormatter(formatter)

    # 给logger添加handler
    logger.addHandler(file_handler)

def login_sap_connect():
    kill_sap()
    sap_app = Config['SAP_app']  # 您的saplogon程序本地完整路径
    # subprocess.Popen([sap_app])
    # 使用 win32api.ShellExecute 启动 SAP Logon
    win32api.ShellExecute(0, 'open', sap_app, '', '', 1)
    time.sleep(1)
    flt = 0
    while flt == 0:
        try:
            hwnd = win32gui.FindWindow(None, Config['SAP_version'])
            window_placement = win32gui.GetWindowPlacement(hwnd)
            if window_placement[1] == win32con.SW_SHOWMINIMIZED:
                # 还原窗口
                win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
                print('已经还原窗口')
            else:
                # 如果窗口已经正常显示，则激活窗口
                win32gui.SetForegroundWindow(hwnd)
            flt = win32gui.FindWindowEx(hwnd, None, "Edit", None)  # capture handle of filter
        except:
            time.sleep(0.5)
    print(flt)
    logger.info(flt)
    print(hwnd)
    logger.info(hwnd)
    win32gui.SendMessage(flt, win32con.WM_SETTEXT, None, Config["SAP_SID"])
    win32gui.SendMessage(flt, win32con.WM_KEYDOWN, win32con.VK_RIGHT, 0)
    win32gui.SendMessage(flt, win32con.WM_KEYUP, win32con.VK_RIGHT, 0)
    time.sleep(2)
    # 单击Button登录
    dlg = win32gui.FindWindowEx(hwnd, None, "Button", None)  # 登陆（0）
    print(dlg)
    logger.info(dlg)
    win32gui.SendMessage(dlg, win32con.WM_LBUTTONDOWN, 0)
    win32gui.SendMessage(dlg, win32con.WM_LBUTTONUP, 0)
    SapGuiAuto = win32com.client.GetObject("SAPGUI")
    if not type(SapGuiAuto) == win32com.client.CDispatch:
        return
    application = SapGuiAuto.GetScriptingEngine
    if not type(application) == win32com.client.CDispatch:
        SapGuiAuto = None
        return
    connection = application.Children(0)
    if not type(connection) == win32com.client.CDispatch:
        application = None
        SapGuiAuto = None
        return
    time.sleep(2)
    flag = 0
    while flag == 0:
        try:
            global session
            session= connection.Children(0)
            flag = 1
        except:
            time.sleep(0.5)
    if not type(session) == win32com.client.CDispatch:
        connection = None
        application = None
        SapGuiAuto = None
        return
    print("SAP Login")
    login_sap()

def login_sap():
    session.findById("wnd[0]").maximize()
    #session.findById('wnd[0]/usr/txtRSYST-MANDT').text ='100'
    session.findById("wnd[0]/usr/txtRSYST-BNAME").text = Config['username']  # SAP登陆用户名
    logger.info('输入账号'+Config['username'])
    session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = Config['SAP_password']  # SAP登陆密码
    logger.info('输入密码')
    session.findById("wnd[0]").sendVKey(0)
    logger.info('登录完成')
    session.findById("wnd[0]/tbar[0]/okcd").text = "mm03"
    session.findById("wnd[0]").sendVKey(0)

def config():
    global Config


    Config={}
    # config的相对路径
    project_root = get_project_root()
    file_path = os.path.join(project_root, '01_Conf', 'Config_download.xlsx')
    df = pd.read_excel(file_path, sheet_name='Variable', header=None, usecols=[0, 1])
    print(df)
    print(df.values)

    for i in range(len(df.values)):
        print(df.values[i][0])
        print(df.values[i][1])
        Config.update({df.values[i][0]: df.values[i][1]})

    print(Config['Name'])
    decode_str=base64.b64encode(Config['Password'].encode())
    print(decode_str)
    encode_str=base64.b64decode(Config['Password']).decode()
    Config['Password']=encode_str
    print(Config['Password'])
    encode_str = base64.b64decode(Config['SAP_password']).decode()
    Config['SAP_password'] = encode_str
    print(Config['SAP_password'])

def download_file(password,username,server):
    #生成文件夹
    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    year = now.strftime("%Y")
    month = now.strftime("%m")
    year_folder_path=os.path.join(Config['Drawing_folder']+"\\po-list",str(year))
    month_folder_path=os.path.join(year_folder_path,str(month))
    daily_folder_path=os.path.join(month_folder_path,str(day))

    #创建年文件夹
    if not os.path.exists(year_folder_path):
        os.makedirs(year_folder_path)
        print(f'Create folder:{year_folder_path}')

    #创建月文件夹
    if not os.path.exists(month_folder_path):
        os.makedirs(month_folder_path)
        print(f'Create folder:{month_folder_path}')

    #创建日文件夹
    if not os.path.exists(daily_folder_path):
        os.makedirs(daily_folder_path)
        print(f'Create folder:{daily_folder_path}')

    #清空所有文件
    def clear_folder(folder_path):
        for filename in os.listdir(folder_path):
            file_path=os.path.join(folder_path, filename)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    print(f'Delete file:{file_path}')
                except Exception as e:
                    print(f'Failed to delete file:{file_path}, error:{e}')

    clear_folder(daily_folder_path)
    # 设置你的Exchange账户和服务器信息
    creds = Credentials(username=username, password=password)
    config = Configuration(server=server, credentials=creds)

    # 创建Account对象
    account = Account(primary_smtp_address=username, config=config, autodiscover=False,
                      access_type=DELEGATE)

    # 获取今天的日期
    today = datetime.now(pytz.utc).date()

    # 创建今天午夜的时区感知的datetime对象
    # 创建今天午夜的CST时区感知的datetime对象
    cst_tz = pytz.timezone('Asia/Shanghai')
    start_time = datetime.combine(today, datetime.min.time()).replace(tzinfo=cst_tz)

    # 创建明天午夜的时区感知的datetime对象
    end_time = start_time + timedelta(days=1)

    # 访问收件箱并获取今天收到的邮件
    inbox = account.inbox
    print(start_time, end_time)
    print(inbox)
    messages = inbox.filter(datetime_received__range=(start_time, end_time))
    print(messages)

    # 下载邮件附件
    for message in messages:
        print(f"Subject: {message.subject}")
        print(f"Received at: {message.datetime_received}")


        for attachment in message.attachments:
            if "po_list" in attachment.name:
                file_path = os.path.join(daily_folder_path, attachment.name)
                with open(file_path, 'wb') as f:
                    f.write(attachment.content)
                print(f"Downloaded attachment: {attachment.name} to {file_path}")

#对下载下来的文件进行修改
def excel_change():
    #先删除第一行
    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    year = now.strftime("%Y")
    month = now.strftime("%m")
    #这是存放po-list的文件
    year_folder_path=os.path.join(Config['Drawing_folder']+"\\po-list",str(year))
    month_folder_path=os.path.join(year_folder_path,str(month))
    daily_folder_path=os.path.join(month_folder_path,str(day))
    wb=load_workbook(daily_folder_path+"\\po_list_lnsc.nopag.xlsx")
    sheet=wb.active
    if sheet.max_row>1:
        sheet.delete_rows(1)
    wb.save(daily_folder_path+"\\po_list_lnsc.nopag.xlsx")
    #这里获取十四天数据
    dataframe_sap=pd.read_excel(daily_folder_path+"\\po_list_lnsc.nopag.xlsx",sheet_name='PO_List')
    print(dataframe_sap)
    # 将 'cof.date' 列转换为 datetime 类型

    dataframe_sap['Conf. Date']=pd.to_datetime(dataframe_sap['Conf. Date'],format='%Y%m%d', errors='coerce')
    print(dataframe_sap['Conf. Date'])
    # 获取今天的日期
    today = datetime.now().date()
    # 计算从今天开始往后十四天的日期范围
    end_date = today + timedelta(days=30)
    # 筛选出从今天开始往后十四天的数据
    dataframe_sap = dataframe_sap[(dataframe_sap['Conf. Date'].dt.date >= today) & (dataframe_sap['Conf. Date'].dt.date <= end_date)]
    print("这是根据conf.date筛选完十四突")
    print(dataframe_sap)
    dataframe_sap=dataframe_sap.drop(columns=['Currency','Unit price','Total Value','Required ','Qty. deliv.','Qty. open','Cost Center','Cost Center Descr.','G/L Account','G/L Account Desc.'])
    #筛选掉值为X的行
    dataframe_sap=dataframe_sap[dataframe_sap['Deliv. completed']!='X']
    print('这是筛掉X的结果')
    print(dataframe_sap)
    # 筛选掉 Deletion Flag 列中值为 L 或 S 的行
    dataframe_sap= dataframe_sap[~dataframe_sap['Deletion Flag'].isin(['L', 'S'])]
    print("这是筛完L或S的结果")
    print(dataframe_sap)
    #去掉3070 1530 1560 1620 1600 1550
    dataframe_sap=dataframe_sap[~dataframe_sap['Mat. Group'].isin(['3070','1530','1600','1560','1620','1550'])]
    #去掉Lechler
    dataframe_sap = dataframe_sap[~dataframe_sap['Supplier Name'].str.contains("echler")]
    print("这是筛掉Lechler的结果,并且是最终结果")
    print(dataframe_sap)
    # 读取 Excel 文件
    file_path = daily_folder_path+"\\po_list_lnsc.nopag.xlsx"
    wb = load_workbook(filename=file_path)

    # 选择或创建工作表
    sheet_name = 'PO_List'
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)

    # 清空工作表
    sheet.delete_rows(1, sheet.max_row)

    # 保存修改后的 Excel 文件
    wb.save(file_path)
    dataframe_sap.to_excel(file_path, sheet_name=sheet_name, index=False)

    destination_file=daily_folder_path+f"\\{day}-polist.xlsx"
    print(f"已将 DataFrame 插入到 {sheet_name} 工作表并保存到 {file_path}")
    os.replace(file_path,destination_file)
    print(f"文件重命名{destination_file}成功")

#sap中获取GuiTableControl (80)数据
def get_data(session, _id, columns=None):
    """
    获取指定列的数据，索引从0开始。

    :param session: SAP的GuiSession对象。
    :param _id: SAP组件ID。
    :param columns: 需要获取数据的列索引。
    """
    result = []
    tbl = session.findById(_id)
    # 如果没有指定要读取的列，则读取全部列
    columns = columns or [i for i in range(tbl.Columns.Count)]

    row_number = 0
    page = 0
    old_position = -1
    new_position = 0

    # 滚动到下页的位置没有变更，则表示到达最后一页
    while old_position != new_position:
        rows_per_page = tbl.Children.Count // tbl.Columns.Count
        page += 1

        # 读取一页的数据
        for row in range(rows_per_page):
            row_data = [tbl.GetCell(row, column_number).Text for column_number in columns]
            result.append(row_data)
            # print("Page:", page, "PageSize:", str(row + 1) + '/' + str(rows_per_page), row_data)
        row_number += rows_per_page

        # 滚动到下页的位置
        next_page_position = tbl.VerticalScrollbar.Position + tbl.VisibleRowCount
        old_position = tbl.verticalScrollbar.position
        tbl.verticalScrollbar.position = next_page_position

        # 获取翻页后的表格控件对象
        tbl = session.findById(_id)
        new_position = tbl.verticalScrollbar.position

    return result
#这是在浏览器上下载图纸的函数
def downlaod_file_chrome(po,material,folder_path,max_version,max_document):
    print(f"这是download_file_chrome方法里的po{po},material{material},folderpath{folder_path}")

    '''
    chrome_options = Options()
    chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")

    # 显式指定 ChromeDriver 路径
    chrome_driver_path = Config['webdriver']

    service = Service(chrome_driver_path)

    # 初始化 WebDriver
    try:
        driver = webdriver.Chrome(service=service, options=chrome_options)
        print("成功连接到 Chrome 浏览器")
    except Exception as e:
        print(f"无法连接到 Chrome 浏览器: {e}")
        exit(1)
    time.sleep(10)
    #点击下载按钮
    #button_dowmnload=driver.find_element(By.XPATH,"/html/body/pdf-viewer//viewer-toolbar//div/div[3]/viewer-download-controls//cr-icon-button//div/cr-icon")


    def wait_for_element_to_appear(driver, xpath):
        while True:
            try:
                # 尝试查找元素
                element = driver.find_element(By.XPATH, xpath)
                # 如果元素存在，退出循环
                break
            except NoSuchElementException:
                # 元素不存在，继续等待
                time.sleep(1)  # 等待 1 秒再检查
    wait_for_element_to_appear(driver, "/html/body/pdf-viewer//viewer-toolbar//div/div[3]/viewer-download-controls//cr-icon-button//div/cr-icon")



    '''
    time.sleep(10)
    def find_latest_pdf(source_dir):
        """
        遍历指定文件夹，找到最新的 PDF 文件。
        :param source_dir: 源文件夹路径
        :return: 最新的 PDF 文件路径
        """
        latest_file = None
        latest_time = None

        for root, dirs, files in os.walk(source_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    file_path = os.path.join(root, file)
                    file_time = os.path.getmtime(file_path)
                    if latest_time is None or file_time > latest_time:
                        latest_time = file_time
                        latest_file = file_path
        return latest_file





        kill_chrome_processes()

    def rename_and_move_pdf(latest_pdf, target_dir, new_name):
        """
        重命名并移动 PDF 文件到目标文件夹。
        :param latest_pdf: 最新的 PDF 文件路径
        :param target_dir: 目标文件夹路径
        :param new_name: 新文件名
        """
        if latest_pdf is None:
            print("没有找到 PDF 文件")
            return

        # 使用目标文件夹名作为PDF文件名（与文件夹保持一致）
        new_filename = f"{os.path.basename(target_dir)}.pdf"

        # 确保目标文件夹存在
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)

        # 构建新文件路径
        new_file_path = os.path.join(target_dir, new_filename)

        # 重命名并移动文件
        shutil.move(latest_pdf, new_file_path)
        print(f"文件已重命名并移动到: {new_file_path}")
        logger.info(f"文件已重命名并移动到: {new_file_path}")

    source_dir=Config['download_default']
    taget_dir=folder_path
    new_name=str(material)+"-"+str(max_version)+"-"+str(max_document)
    def find_and_check_file(base_folder,target_pdf_name,target_folder):
        # 遍历五级文件夹结构
        for root, dirs, files in os.walk(base_folder):
            for file in files:
                if file == target_pdf_name:
                    # 找到目标PDF文件
                    source_pdf_path = os.path.join(root, file)
                    target_pdf_path = os.path.join(target_folder, file)

                    # 复制PDF文件到目标文件夹
                    shutil.copy2(source_pdf_path, target_pdf_path)
                    print(f"找到并复制了PDF文件: {source_pdf_path} -> {target_pdf_path}")
                    return True
        # 如果没有找到目标文件，返回False
        return False
    if find_and_check_file(os.path.join(get_base_path(), '..', '06_DataOutput'),os.path.basename(taget_dir)+'.pdf',taget_dir):
        print("Have found the pdf file")
        logger.info("Have found the pdf file")
    else:
        latest_pdf=find_latest_pdf(source_dir)
        rename_and_move_pdf(latest_pdf, taget_dir, new_name)



def check_version(po,material,version_current):
    found=False

    def check_pdf_files(directory):
        # 列出目录中的所有文件
        for filename in os.listdir(directory):
            if 'po' in filename and 'material' in filename and filename.endswith('.pdf'):
                print(f"匹配到版本以及{po}-{material}一致的{filename}")
                return True  # 一旦匹配到，立即返回 True
        return False  # 遍历完所有文件后，如果没有匹配到，返回 False
    for entry in os.listdir(Config['Drawing_folder']):
        if entry.startswith("20"):
            entry_path=os.path.join(Config['Drawing_folder'],entry)
            # 遍历第一级子文件夹
            for first_level in os.listdir(entry_path):
                first_level_path = os.path.join(entry_path, first_level)
                if os.path.isdir(first_level_path):
                    # 遍历第二级子文件夹
                    for second_level in os.listdir(first_level_path):               
                        second_level_path = os.path.join(first_level_path, second_level)
                        if os.path.isdir(second_level_path):
                            # 遍历第三级子文件夹
                            for third_level in os.listdir(second_level_path):
                                third_level_path = os.path.join(second_level_path, third_level)
                                if os.path.isdir(third_level_path):
                                    # 检查文件夹名称是否包含 'po' 和 'material'
                                    if str(po) in third_level and str(material) in third_level:
                                        # 进一步判断 version
                                        # 假设 version 是文件夹名称的一部分
                                        version = third_level.split('-')[-1]  # 假设版本号在名称的最后部分
                                        print(f"Found folder: {third_level_path}, Version: {version}")
                                        logger.info(f"Found folder: {third_level_path}, Version: {version}")
                                        if not version.isdigit() or not version_current.isdigit():
                                            found= True
                                        else:
                                            # 将 version 和 version_current 转换为整数
                                            version_int = int(version)
                                            version_current_int = int(version_current)

                                            # 比较版本号
                                            if version_current_int > version_int:
                                                print("版本号不一致，需要更新")
                                                found = False
                                            elif version_current_int == version_int:
                                                print("版本号一致，不需要更新")
                                                found = True
                                            else:
                                                found=True
    return not found









#获取文档各种语言描述txt的方法
def gettxt(file_path):
    session.findById("wnd[0]/tbar[1]/btn[27]").press()
    session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP14").select()

    def exists(element_path):
        try:
            # 尝试获取元素
            element = session.FindById(element_path)
            if element:
                print("元素存在")
                return True
            else:
                print("元素不存在")
                return False
        except Exception as e:
            print(f"元素不存在或发生其他错误: {e}")
            return False

    # 创建空的 DataFrame 并指定列名
    df = pd.DataFrame(columns=['name', 'description'])

    for i in range(1, 6):
        print(f"检查第{i}行,这是元素 wnd[0]/usr/tabsTABSPR1/tabpSP14/ssubTABFRA1:SAPLMGMM:2010/subSUB2:SAPLMGD1:2321/tblSAPLMGD1TC_LONGTEXT/btnSELE[0,{str(i-1)}]")
        if exists(f"wnd[0]/usr/tabsTABSPR1/tabpSP14/ssubTABFRA1:SAPLMGMM:2010/subSUB2:SAPLMGD1:2321/tblSAPLMGD1TC_LONGTEXT/btnSELE[0,{str(i-1)}]"):

            print(f"第{i}行存在")

            session.findById(f"wnd[0]/usr/tabsTABSPR1/tabpSP14/ssubTABFRA1:SAPLMGMM:2010/subSUB2:SAPLMGD1:2321/tblSAPLMGD1TC_LONGTEXT/btnSELE[0,{str(i-1)}]").press()
            text=session.findById(f"wnd[0]/usr/tabsTABSPR1/tabpSP14/ssubTABFRA1:SAPLMGMM:2010/subSUB2:SAPLMGD1:2321/tblSAPLMGD1TC_LONGTEXT/txtLANG_TC_TAB_TC-SPTXT[1,{i-1}]").text
            print(f'这是第{i}行的{text}')
            text_sheel=session.findById(f"wnd[0]/usr/tabsTABSPR1/tabpSP14/ssubTABFRA1:SAPLMGMM:2010/subSUB2:SAPLMGD1:2321/cntlLONGTEXT_BESTELL/shellcont/shell").text
            if ('___________') not in text:
                newrow={"name":text,"description":text_sheel}
                df.loc[len(df)] = newrow

    print("这是要插入到txt里的数据")
    print(df)
    # 定义要删除的模式
    pattern = r'^____________$'

    # 使用布尔索引删除符合条件的行
    df = df[~df['name'].str.contains(pattern, regex=True)]
    print("这是最终要插入到txt里的数据")
    print(df)
    # 将 DataFrame 写入文件，追加模式
    #df.to_csv(file_path, mode='a', header=not pd.io.common.file_exists(file_path), index=False, sep='\t')
    #这里添加换行符，确保不会紧挨着
    with open(file_path, 'a',encoding='utf-8') as f:
        for _, row in df.iterrows():
            f.write(f"{row['name']}:\n{row['description']}\n\n")

    #这个方法不用
def sap_control(po,material,desc):
    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    year = now.strftime("%Y")
    month = now.strftime("%m")
    year_folder_path=os.path.join(Config['Drawing_folder'],str(year))
    month_folder_path=os.path.join(year_folder_path,str(month))
    daily_folder_path=os.path.join(month_folder_path,str(day))
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").text = material
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").caretPosition = 6
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]").sendVKey(0)
    # 将焦点设置到标签页
    session.findById("wnd[0]/usr/tabsTABSPR1").setFocus()
    session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP26").select()
    checkbox = session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP26/ssubTABFRA1:SAPLMGMM:2000/subSUB2:SAPLMGD1:2751/chkMARC-QMATV")

    # 判断复选框是否被选中
    if checkbox.selected:
        print("复选框已被选中，需要继续运行")

        session.findById("wnd[0]/tbar[1]/btn[30]").press()
        # 选择标签页
        session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU01").select()
        print(f"控件类型：{session.findById('wnd[0]/usr/tabsTABSPR1/tabpZU01')}")
        # 获取表格控件
        path = "wnd[0]/usr/tabsTABSPR1/tabpZU01/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLZ_MMGD1:8000/tblSAPLZ_MMGD1TC_KTXT"
        table_control = session.findById(path)

        # 检查控件类型
        print(f"table_control 类型: {type(table_control)}")

        # 获取表格的行数和列数
        try:
            row_count = table_control.RowCount
            print(f"行数: {row_count}")
        except AttributeError as e:
            print(f"获取行数时发生错误: {e}")
            exit(1)

        # 手动指定列标题
        column_titles = ["Language", "Material"]


        data = get_data(session, path, columns=[0, 1])
        print(data)
        df_description=pd.DataFrame(data,columns=['Lanuage','Description'])

        def remove_illegal_chars(s):
            illegal_chars = r'*?<>|:/\\'
            return ''.join(c for c in s if c not in illegal_chars)
        desc=remove_illegal_chars(desc)
        #将数据写入txt
        output_file=daily_folder_path+"\\"+f"{po}-{material}-{desc}.txt"
        df_description.to_csv(output_file,sep='\t',index=False)
        print("各种语言描述插入成功")
        #切换到下载文件界面
        session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04").select()
        radio_button=session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subBUTTON:SAPLCV140:0203/radGF_ALL_REL")
        radio_button.selected=True
        print("All Released Version已经设置为选中状态")
        time.sleep(1)
        #text=session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").text
        #print(text)
        state=[]
        version=[]
        document=[]
        description=[]
        for i in range(session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").RowCount):
            print(i)
            rowValueTcode=session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                i, 'DOKAR')
            state.append(rowValueTcode)
            versionCode = session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                i, 'DOKVR')
            version.append(versionCode)
            documentCode = session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                i, 'DOKNR')
            document.append(documentCode)
            descriptionCode = session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                i, 'DKTXT')
            description.append(descriptionCode)
        print(state)
        print(version)
        print(document)
        print(description)
        df_data=pd.DataFrame({
            'Type':state,
            'Version':version,
            'Document':document,
            'Description':description
        })
        print(df_data)
        df_data=df_data[df_data['Type']=='ZDR']
        if not df_data.empty:
            print("没有找到ZDR类型的数据")
            df_data = df_data.sort_values('Version', ascending=False)
            max_version_row = df_data.iloc[0]
            max_version = max_version_row['Version']
            print(f"这是最大版本{max_version}")
            max_index = df_data.index[0]
            print(f"这是最大版本索引号{max_index}")
            session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").SelectedRows = max_index
            session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").pressToolbarButton(
                "ICON_DISPLAY")

            def element_exists(session, element_id):
                try:
                    # 尝试查找元素
                    element = session.findById(element_id)
                    print(f"Element '{element_id}' exists.")
                    return True
                except pywintypes.com_error as e:
                    # 如果抛出 com_error 异常，说明元素不存在
                    print(f"Element '{element_id}' does not exist. Error: {e}")
                    return False

            exists = element_exists(session, "wnd[1]/usr/txtMESSTXT1")
            if exists:
                print("元素存在，没有权限下载图纸")
                session.findById("wnd[1]/tbar[0]/btn[0]").press()

            else:
                print("有权限，这里继续下载图纸")
        session.findById("wnd[0]/tbar[1]/btn[17]").press()
    else:
        session.findById("wnd[0]/tbar[1]/btn[17]").press()

#这是生成txt和excel文件的路径
def txt_and_excel(folder_path,desc,po,material,version):
    #session.findById("wnd[0]/tbar[1]/btn[27]").press()
    # 选择标签页
    session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU01").select()
    print(f"控件类型：{session.findById('wnd[0]/usr/tabsTABSPR1/tabpZU01')}")
    # 获取表格控件
    path = "wnd[0]/usr/tabsTABSPR1/tabpZU01/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLZ_MMGD1:8000/tblSAPLZ_MMGD1TC_KTXT"
    table_control = session.findById(path)

    # 检查控件类型
    print(f"table_control 类型: {type(table_control)}")

    # 获取表格的行数和列数
    try:
        row_count = table_control.RowCount
        print(f"行数: {row_count}")
    except AttributeError as e:
        print(f"获取行数时发生错误: {e}")
        exit(1)


    data = get_data(session, path, columns=[0, 1])
    print(data)
    df_description = pd.DataFrame(data, columns=['Lanuage', 'Description'])



    def remove_illegal_chars(s):
        illegal_chars = r'*?<>|:/\\'
        return ''.join(c for c in s if c not in illegal_chars)

    desc = remove_illegal_chars(desc)
    # 将数据写入txt
    output_file = folder_path + "\\" + f"{po}-{material}-{desc}.txt"
    #df_description.to_csv(output_file, sep='\t', index=False)
    #print("各种语言描述插入成功")
    # 打开文件并写入三个字符串
    with open(output_file, 'a', encoding='utf-8') as file:
        file.write(str(po) + '\n')
        file.write(str(material) + '\n')
        file.write(str(desc)+ '\n')



    def find_and_check_file(base_folder, target_pdf_name, target_folder):
        # 遍历五级文件夹结构
        for root, dirs, files in os.walk(base_folder):
            for file in files:
                if file == target_pdf_name:
                    # 找到目标PDF文件
                    source_pdf_path = os.path.join(root, file)
                    target_pdf_path = os.path.join(target_folder, file)

                    # 复制PDF文件到目标文件夹
                    shutil.copy2(source_pdf_path, target_pdf_path)
                    print(f"找到并复制了PDF文件: {source_pdf_path} -> {target_pdf_path}")
                    return True
        # 如果没有找到目标文件，返回False
        return False

    source_path=os.path.join(Config['Folder_index'],"FM-QA-018 B 进料检验记录.xlsx")
    destination_path=os.path.join(folder_path,"FM-QA-018 B 进料检验记录.xlsx")
    if find_and_check_file(Config['Folder_index'], material+'-'+version+'.xlsx', folder_path):
        print("The file is exist and copy successfully")
        logger.info("The file is exist and copy successfully")
    else:
        # 检查源文件是否存在
        if os.path.exists(source_path):
            # 复制文件
            shutil.copy2(source_path, destination_path)
            print(f"File copied successfully from {source_path} to {destination_path}")
        else:
            print(f"Source file {source_path} does not exist")

    gettxt(output_file)




def element_exists(session, element_id):
    try:
        # 尝试查找元素
        element = session.findById(element_id)
        print(f"Element '{element_id}' exists.")
        logger.info(f"Element '{element_id}' exists.")
        return True
    except pywintypes.com_error as e:
        # 如果抛出 com_error 异常，说明元素不存在
        print(f"Element '{element_id}' does not exist. Error: {e}")
        logger.info(f"Element '{element_id}' does not exist. Error: {e}")
        return False

def send_email(subject, body, to_email, username, password, exchange_server, email_address,attachment_path):
    # 创建Credentials对象
    creds = Credentials(username=username, password=password)

    # 创建Configuration对象
    config = Configuration(server=exchange_server, credentials=creds)

    # 创建Account对象
    account = Account(primary_smtp_address=email_address, config=config, access_type=DELEGATE)

    # 创建Message对象并发送邮件
    subject = subject
    body = body
    recipients = to_email

    attachments=[]
    for path in attachment_path:
        file_name = os.path.basename(path)  # ✅ 自动提取文件名
        with open(path, 'rb') as f:
            attachment = FileAttachment(name=file_name, content=f.read())
            attachments.append(attachment)

    # 过滤收件人列表，排除特定的收件人
    excluded_emails = ['a', 'd', 'e', 'i', 'm', 'p','A','D','E','I','M','P'],

    filtered_recipients = [
        r.strip() for r in recipients
        if (r.strip().split('@')[0] if '@' in r.strip() else r.strip()) not in excluded_emails
    ]
    # 创建邮件
    msg = Message(
        account=account,
        subject=subject,
        body=body,

        to_recipients=filtered_recipients,
        attachments=attachments
    )
    # 发送邮件
    msg.send()

def purchcase_receiver():
    receiver = Config['Purchase_Receiver'].rstrip(';').split(';')
    print(receiver)
    return receiver

def sap_control_new(po,material,desc,row,index):
    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    year = now.strftime("%Y")
    month = now.strftime("%m")
    year_folder_path = os.path.join(Config['Drawing_folder'], str(year))
    month_folder_path = os.path.join(year_folder_path, str(month))
    daily_folder_path = os.path.join(month_folder_path, str(day))
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").text = material
    session.findById("wnd[0]/usr/ctxtRMMG1-MATNR").caretPosition = 6
    session.findById("wnd[0]").sendVKey(0)
    session.findById("wnd[1]").sendVKey(0)
    session.findById("wnd[1]").sendVKey(0)
    # 将焦点设置到标签页
    session.findById("wnd[0]/usr/tabsTABSPR1").setFocus()
    if not element_exists(session,"wnd[0]/usr/tabsTABSPR1/tabpSP15"):
        print("不存在SP15页面")
        df_polist.at[index, 'P-S Mat. Status'] = ""
        return False
    session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP15").select()
    status=session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP15/ssubTABFRA1:SAPLMGMM:2000/subSUB2:SAPLZ_MMGD1:2481/ctxtMARC-MMSTA").text
    print(f"status is {status}")
    df_polist.at[index,'P-S Mat. Status'] = status

    if not element_exists(session, "wnd[0]/usr/tabsTABSPR1/tabpSP26"):
        print(f"不存在SP26页面")
        df_polist.at[index, '下载状态'] = "未维护QM"
        df_polist.at[index, '是否是Z01P'] = "未维护QM"
        session.findById("wnd[0]/tbar[0]/btn[3]").press()
        return False

    session.findById("wnd[0]/usr/tabsTABSPR1/tabpSP26").select()
    def exists(element_path):
        try:
            # 尝试获取元素
            element = session.FindById(element_path)
            if element:
                print("元素存在")
                return True
            else:
                print("元素不存在")
                return False
        except Exception as e:
            print(f"元素不存在或发生其他错误: {e}")
        return False



    #这里加上判断是否为Z01P的逻辑


    #session.findById("wnd[0]/tbar[1]/btn[17]").press()



    
    if exists("wnd[0]/usr/tabsTABSPR1/tabpSP26/ssubTABFRA1:SAPLMGMM:2000/subSUB2:SAPLMGD1:2751/btnMARC_QMPD"):

        checkbox = session.findById(
        "wnd[0]/usr/tabsTABSPR1/tabpSP26/ssubTABFRA1:SAPLMGMM:2000/subSUB2:SAPLMGD1:2751/chkMARC-QMATV")

        # 判断复选框是否被选中
        if checkbox.selected:
            session.findById(
                "wnd[0]/usr/tabsTABSPR1/tabpSP26/ssubTABFRA1:SAPLMGMM:2000/subSUB2:SAPLMGD1:2751/btnMARC_QMPD").press()

            checkbox_01p = None
            checkbox_02p = None

            # 获取表格总行数
            total_rows = session.findById("wnd[1]/usr/tblSAPLQPLSPRUEFDAT").RowCount
            # 获取当前可见行数
            visible_rows = session.findById("wnd[1]/usr/tblSAPLQPLSPRUEFDAT").VisibleRowCount
            print('总行数' + str(total_rows))
            logger.info('总行数' + str(total_rows))
            print('可见行数' + str(visible_rows))
            logger.info('可见行数' + str(visible_rows))

            # 遍历判断是否是Z01P
            for i in range(0, int(total_rows) - 1):
                try:
                    print(f"/app/con[0]/ses[0]/wnd[1]/usr/tblSAPLQPLSPRUEFDAT/ctxtRMQAM-ART[1,{i}]")
                    logger.info(f"/app/con[0]/ses[0]/wnd[1]/usr/tblSAPLQPLSPRUEFDAT/ctxtRMQAM-ART[1,{i}]")
                    text = session.findById(f"wnd[1]/usr/tblSAPLQPLSPRUEFDAT/ctxtRMQAM-ART[1,{i}]").text
                    logger.info(f'第{i}行的text为' + text)

                    # print(session.findById(f"wnd[1]/usr/tblSAPLQPLSPRUEFDAT/ctxtRMQAM-ART[1,0]").value)
                    if str(text) == "Z01P":
                        checkbox_01p = f"wnd[1]/usr/tblSAPLQPLSPRUEFDAT/chkRMQAM-APA[3,{i}]"
                        checkbox_02p = f"wnd[1]/usr/tblSAPLQPLSPRUEFDAT/chkRMQAM-AKTIV[4,{i}]"
                        print(f'Z01P在第{i}行')
                        break
                except Exception as e:
                    print(f"获取第{i}行数据时发生错误: {e}")
                    logger.info(f"获取第{i}行数据时发生错误: {e}")
                    break

            if exists(checkbox_01p):
                if session.findById(checkbox_01p).selected == True and session.findById(checkbox_02p).selected == True:
                    print("01p复选框和02p复选框均已被选中")
                    logger.info("01p复选框和02p复选框均已被选中")
                    df_polist.at[index, '是否是Z01P'] = "激活状态"
                else:
                    print("01p复选框和02p复选框未被选中")
                    logger.info("01p复选框和02p复选框未被选中")
                    df_polist.at[index, '是否是Z01P'] = "非激活状态"

                session.findById("wnd[1]/tbar[0]/btn[0]").press()
            else:
                df_polist.at[index, '是否是Z01P'] = "非激活状态"
                print("未找到01p复选框和02p复选框")
                logger.info("未找到01p复选框和02p复选框")

            print(df_polist)

            if exists('wnd[1]/tbar[0]/btn[0]'):
                session.findById("wnd[1]/tbar[0]/btn[0]").press()
            print(f"{po}+{material}复选框已被选中，需要继续运行")
            logger.info(f"{po}+{material}复选框已被选中，需要继续运行")
            #那么需要先到图纸界面查看是否有图纸，然后生成文件夹
            # 切换到下载文件界面
            session.findById("wnd[0]/tbar[1]/btn[30]").press()
            session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04").select()
            radio_button=session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subBUTTON:SAPLCV140:0203/radGF_ALL_REL")
            radio_button.selected=True
            print("All Released Version已经设置为选中状态")
            time.sleep(1)

            state = []
            version = []
            document = []
            description = []
            for i in range(session.findById(
                    "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").RowCount):
                print(i)
                rowValueTcode = session.findById(
                    "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                    i, 'DOKAR')
                state.append(rowValueTcode)
                versionCode = session.findById(
                    "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                    i, 'DOKVR')
                version.append(versionCode)
                documentCode = session.findById(
                    "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                    i, 'DOKNR')
                document.append(documentCode)
                descriptionCode = session.findById(
                    "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").GetCellValue(
                    i, 'DKTXT')
                description.append(descriptionCode)
            print(state)
            print(version)
            print(document)
            print(description)
            df_data = pd.DataFrame({
                'Type': state,
                'Version': version,
                'Document': document,
                'Description': description
            })
            print(df_data)
            df_data = df_data[df_data['Type'] == 'ZDR']
            if not df_data.empty:
                print("找到ZDR类型的数据")
                logger.info("找到ZDR类型的数据")

                df_data = df_data.sort_values('Version', ascending=False)
                max_version_row = df_data.iloc[0]
                max_version = max_version_row['Version']
                max_document = max_version_row['Document']
                print('这是最大版本行的document')
                print(f"这是最大版本{max_version}")
                logger.info(f"这是最大版本{max_version}")
                #找到ZDR类型的数据，那么根据最大版本进行匹配，如果之前没有比该版本更大的文件，那么下载
                '''
                
                中间还有一个检查版本号的方法没有写
                
                
                
                '''

                if check_version(po,material,max_version):
                    print("没有比该版本大的文件，可以下载")
                    logger.info("没有比该版本大的文件，可以下载")



                    max_index = df_data.index[0]
                    print(f"这是最大版本索引号{max_index}")
                    logger.info(f"这是最大版本索引号{max_index}")
                    session.findById(
                        "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").SelectedRows = max_index
                    session.findById(
                        "wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").pressToolbarButton(
                        "ICON_DISPLAY")


                    exists = element_exists(session, "wnd[1]/usr/txtMESSTXT1")
                    if exists:
                        print("元素存在，没有权限下载图纸")
                        logger.info("元素存在,没有权限下载图纸")
                        session.findById("wnd[1]/tbar[0]/btn[0]").press()
                        #不管有没有权限都要获取txt以及excel
                        # 生成文件夹
                        folder_path = os.path.join(daily_folder_path,
                                                str(po) + "-" + str(material) + "-NA")
                        # 创建文件文件夹
                        if not os.path.exists(folder_path):
                            os.makedirs(folder_path)
                            print(f'Create folder:{folder_path}')
                            logger.info(f'Create folder:{folder_path}')
                        #row['下载状态']='无权限，E:000'
                        df_polist.at[index,'下载状态']='无权限，E:000'
                        print(f"设置{po}+{material}下载状态为{row['下载状态']}")
                        logger.info(f"设置{po}+{material}下载状态为{row['下载状态']}")

                    else:
                        print("有权限，这里继续下载图纸")
                        logger.info("有权限，这里继续下图纸")
                        folder_path = os.path.join(daily_folder_path,
                                                str(po) + "-" + str(material) + "-"+str(max_version))
                        # 创建文件文件夹
                        if not os.path.exists(folder_path):
                            os.makedirs(folder_path)
                            print(f'Create folder:{folder_path}')
                            logger.info(f'Create folder:{folder_path}')
                        #session.findById("wnd[0]/usr/tabsTABSPR1/tabpZU04/ssubTABFRA1:SAPLMGMM:2110/subSUB2:SAPLMGD1:3400/subDOCU:SAPLCV140:0204/subDOC_ALV:SAPLCV140:0206/cntlALV_CUST_DOC/shellcont/shell").pressToolbarButton("ICON_DISPLAY")
                        downlaod_file_chrome(po,material,folder_path,max_version,max_document)
                        update_summary_table(po, material, max_version, max_document, folder_path)
                        #row['下载状态'] = "已更新，已下载"
                        df_polist.at[index, '下载状态'] = "已更新，已下载"
                        print(f"设置{po}+{material}下载状态为{row['下载状态']}")
                        logger.info(f"设置{po}+{material}下载状态为{row['下载状态']}")

                    #通过check方法判断了需要生成文件夹，以下是生成txt以及复制excel文件的步骤
                    txt_and_excel(folder_path,desc,po,material,str(max_version))


                else:
                    print("该文件版本没有更新，不用下载")
                    logger.info("该文件版本没有新增，不用下载")
                    #row['下载状态']="未更新，无需下载"
                    df_polist.at[index, '下载状态'] ="未更新，无需下载"
                    print(f"设置{po}+{material}下载状态为{row['下载状态']}")
                    logger.info(f"设置{po}+{material}下载状态为{row['下载状态']}")
            else:
                print("没有图纸，只生成空文件夹")
                logger.info("没有图纸，只生成空文件夹")
                folder_path = os.path.join(daily_folder_path, str(po) + "-" + str(material) + "-NA")
                print(f"创建文件夹{folder_path}")
                if check_version(po,material, "NA"):
                    # 创建文件文件夹
                    if not os.path.exists(folder_path):
                        os.makedirs(folder_path)
                        print(f'Create folder:{folder_path}')
                        logger.info(f'Create folder:{folder_path}')
                    txt_and_excel(folder_path, desc, po, material,'NA')
                    df_polist.at[index, '下载状态'] ="已更新，已下载,无图纸"
                    #row['下载状态']="已更新，已下载,无图纸"
                    print(f"设置{po}+{material}下载状态为{row['下载状态']}")
                    logger.info(f"设置{po}+{material}下载状态为{row['下载状态']}")

                else:
                    df_polist.at[index, '下载状态'] = "没有图纸且已经生成过文件夹"
                    print(f"{po}-{material}没有图纸且已经生成过文件夹")
                    logger.info(f"{po}+{material}没有图纸且已经生成过文件夹")

            session.findById("wnd[0]/tbar[1]/btn[17]").press()
        else:
            print(f"{po}+{material}复选框未被选中，直接运行下一条")
            logger.info(f"{po}+{material}复选框未被选中，直接运行下一条")
            #row['下载状态']="未设Q,无需下载"
            df_polist.at[index, '下载状态'] ="未设Q,无需下载"
            print(f"设置下载状态未{row['下载状态']}")
            logger.info(f"设置下载状态未{row['下载状态']}")
            session.findById("wnd[0]/tbar[1]/btn[17]").press()
    else:
        print(f"{po}+{material}不存在Insp Setup按钮")
        logger.info(f"{po}+{material}不存在Insp Setup按钮")
        #row['下载状态']="未设Q,无需下载"
        df_polist.at[index, '下载状态'] ="未设Q,无需下载"
        print(f"设置下载状态未{row['下载状态']}")
        logger.info(f"设置下载状态未{row['下载状态']}")
        while True:
            if exists("wnd[0]/usr/ctxtRMMG1-MATNR"):
               
                break
            session.findById("wnd[0]/tbar[1]/btn[17]").press()
    



    return True


#这里防止出错，单独写一个方法用来遍历数据
def update_summary_table(po, material, version, document, folder_path):
    summary_file = os.path.join(Config['Drawing_folder'], 'drawing_summary.xlsx')
    summary_dir = Config['Drawing_folder']
    key = f"{po}-{material}-{version}"
    pdf_name = f"{po}-{material}-{version}.pdf"
    pdf_abs_path = os.path.join(folder_path, pdf_name)
    try:
        rel_path = os.path.relpath(pdf_abs_path, summary_dir)
    except ValueError:
        rel_path = pdf_abs_path
    today_str = datetime.now().strftime('%Y-%m-%d')

    if os.path.exists(summary_file):
        wb = load_workbook(summary_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(['PO-Material-\u7248\u672c', '\u4e0b\u8f7d\u7684\u56fe\u7eb8\u8def\u5f84', '\u4e0b\u8f7d\u7684\u65e5\u671f'])

    for existing_row in ws.iter_rows(min_row=2, values_only=True):
        if existing_row[0] == key:
            print(f"[summary] {key} already exists, skip")
            logger.info(f"[summary] {key} already exists, skip")
            return

    new_row_idx = ws.max_row + 1
    ws.cell(row=new_row_idx, column=1, value=key)
    link_cell = ws.cell(row=new_row_idx, column=2, value=rel_path)
    link_cell.hyperlink = rel_path
    link_cell.font = Font(color="0563C1", underline="single")
    ws.cell(row=new_row_idx, column=3, value=today_str)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 15
    wb.save(summary_file)
    print(f"[summary] appended: {key} -> {rel_path}")
    logger.info(f"[summary] appended: {key} -> {rel_path}")


def read_excel():

    def login_sap_connect_get(retry_count=7):
        attempts=0
        while attempts<retry_count:
            try:
                login_sap_connect()
                break
            except Exception as e:
                attempts+=1
                logger.error(f"登录失败{e}")
                if attempts<retry_count:
                    logger.info(f"重试 ({attempts}/{retry_count})...")
                else:
                    logger.error(f"达到最大登录次数，登录失败{e}")

    login_sap_connect_get()


    now = datetime.now()
    day = now.strftime("%Y-%m-%d")
    year = now.strftime("%Y")
    month = now.strftime("%m")
    #这是存放po-list的文件
    year_folder_path=os.path.join(Config['Drawing_folder'],str(year))
    month_folder_path=os.path.join(year_folder_path,str(month))
    daily_folder_path=os.path.join(month_folder_path,str(day))
    #创建年文件夹,这里创建output里的直属文件夹
    if not os.path.exists(year_folder_path):
        os.makedirs(year_folder_path)
        print(f'Create folder:{year_folder_path}')

    #创建月文件夹
    if not os.path.exists(month_folder_path):
        os.makedirs(month_folder_path)
        print(f'Create folder:{month_folder_path}')

    #创建日文件夹
    if not os.path.exists(daily_folder_path):
        os.makedirs(daily_folder_path)
        print(f'Create folder:{daily_folder_path}')
    year_folder_path=os.path.join(Config['Drawing_folder']+"\\po-list",str(year))
    month_folder_path=os.path.join(year_folder_path,str(month))
    daily_folder_path=os.path.join(month_folder_path,str(day))
    destination_file = daily_folder_path + f"\\{day}-polist.xlsx"
    global df_polist
    df_polist=pd.read_excel(destination_file,sheet_name='PO_List')
    print("这是polist的dataframe")
    logger.info("这是polist的是dataframe")
    print(df_polist)
    logger.info(df_polist)
    if '下载状态' not in df_polist.columns:
        df_polist['下载状态'] = ''
    if '是否是Z01P' not in df_polist.columns:
        df_polist['是否是Z01P'] = ''
    for index,row in df_polist.iterrows():
        if not pd.isna(row['SAP Material No.']):
            print(f"开始下载{row['PO No.']}+{str(row['SAP Material No.']).split('.')[0]}")
            logger.info(f"开始下载{row['PO No.']}+{str(row['SAP Material No.']).split('.')[0]}")
            # 遍历每一行数据
            if not sap_control_new(str(row['PO No.']), str(row['SAP Material No.']).split('.')[0], row['Desc.'], row,index):
                continue
        else:
            print(f"{row['PO No.']}+{row['SAP Material No.']}为空，跳过")
            logger.info(f"{row['PO No.']}+{row['SAP Material No.']}为空，跳过")


    # 筛选 'Download Status' 不包含 '已下载' 的行
    print(df_polist)
    logger.info(df_polist)
    print("这是最终结果，所有当天生成的文件夹")
    logger.info("这是最终结果，所有当天生成的文件夹")
    print(df_polist)
    logger.info(df_polist)
    df_polist.to_excel(destination_file,index=False,sheet_name='PO_List')
    print(purchcase_receiver())
    print('文件路径'+str(destination_file))
    send_email(subject='Z01P状态维护', body='Z01P状态维护',
               to_email=purchcase_receiver(), username=Config['Email_name'],
               password=Config['Password'], exchange_server=Config['Email_server'],
               email_address=Config['Email_name'], attachment_path=[rf'{destination_file}'])




def receiver():
    receiver = Config['Receiver'].rstrip(';').split(';')
    print(receiver)
    return receiver



def main():
    try:
        print(f"程序启动，基础路径: {get_base_path()}")
        ensure_working_dir()
        kill_excel_processes()
        #kill_chrome_processes()
        kill_sap()
        config()

        logging_info()
        download_file(Config['Password'], Config['Email_name'], Config['Email_server'])
        excel_change()
        read_excel()
        kill_sap()

        send_email(subject='RPACN01_07'+str(datetime.now())+'质量下载图纸运行完成', body=Config['Body_name_end'],
                   to_email=receiver(), username=Config['Email_name'],
                   password=Config['Password'], exchange_server=Config['Email_server'],
                   email_address=Config['Email_name'],attachment_path='')
    except Exception as e:
        print(f"程序出现异常，请检查日志文件，异常信息为：{e}")
        traceback.print_exc()
        try:
            logger.error(f"程序出现异常，请检查日志文件，异常信息为：{e}")
            logger.error(traceback.format_exc())
        except:
            pass
        input("按回车键退出...")



if __name__ == "__main__":
        main()